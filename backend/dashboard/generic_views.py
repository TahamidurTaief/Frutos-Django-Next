"""
Generic CRUD Views - Universal Dashboard Views for All Models
Auto-generates CRUD operations using model introspection
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.db.models import Q
from django.db import models
from django.contrib import messages
from django.utils.text import slugify
from django.forms import modelform_factory
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.apps import apps
import csv
import json
from datetime import datetime

from .utils import (
    model_registry, get_field_display_value, get_model_fields_for_list,
    get_search_fields, get_filter_fields, check_model_permission, get_model_actions
)


def staff_required(user):
    """Check if user is staff"""
    return user.is_staff


@login_required
@user_passes_test(staff_required)
def generic_model_list(request, app_label, model_name):
    """
    Generic list view for any model
    Features: pagination, search, filters, sorting
    """
    # Get model
    model = apps.get_model(app_label, model_name)
    if not model:
        messages.error(request, f"Model {app_label}.{model_name} not found")
        return redirect('dashboard:home')
    
    # Check permission
    if not check_model_permission(request.user, model, 'view'):
        return HttpResponseForbidden("You don't have permission to view this model")
    
    # Get metadata
    meta = model_registry.get_model_meta(app_label, model_name)
    admin_config = meta.get('admin_config') if meta else None
    
    # Get query parameters
    query = request.GET.get('q', '')
    page_number = request.GET.get('page', 1)
    page_size = int(request.GET.get('page_size', 25))
    sort_by = request.GET.get('sort', '-pk')
    
    # Build queryset
    queryset = model.objects.all()
    
    # Apply search
    if query:
        search_fields = get_search_fields(model, admin_config)
        q_objects = Q()
        for field_name in search_fields:
            q_objects |= Q(**{f"{field_name}__icontains": query})
        queryset = queryset.filter(q_objects)
    
    # Apply filters
    for key, value in request.GET.items():
        if key.startswith('filter_') and value:
            field_name = key.replace('filter_', '')
            try:
                # Boolean filters
                if value.lower() in ['true', 'false']:
                    queryset = queryset.filter(**{field_name: value.lower() == 'true'})
                else:
                    queryset = queryset.filter(**{field_name: value})
            except:
                pass
    
    # Apply sorting
    try:
        queryset = queryset.order_by(sort_by)
    except:
        queryset = queryset.order_by('-pk')
    
    # Optimize query with select_related for FK
    for field in model._meta.get_fields():
        if field.many_to_one and field.concrete:
            try:
                queryset = queryset.select_related(field.name)
            except:
                pass
    
    # Pagination
    paginator = Paginator(queryset, page_size)
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.get_page(1)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)
    
    # Get fields for display
    list_fields = get_model_fields_for_list(model, admin_config)
    filter_fields = get_filter_fields(model, admin_config)
    
    # Get actions available to user
    actions = get_model_actions(request.user, model)
    
    context = {
        'app_label': app_label,
        'model_name': model_name,
        'model': model,
        'meta': meta,
        'objects': page_obj,
        'page_obj': page_obj,
        'list_fields': list_fields,
        'filter_fields': filter_fields,
        'query': query,
        'sort_by': sort_by,
        'actions': actions,
        'page_size': page_size,
    }
    
    # HTMX request - return table partial
    if request.headers.get('HX-Request'):
        return render(request, 'dashboard/generic/partials/table.html', context)
    
    # Full page
    return render(request, 'dashboard/generic/list.html', context)


@login_required
@user_passes_test(staff_required)
def generic_model_create(request, app_label, model_name):
    """Generic create view for any model"""
    model = apps.get_model(app_label, model_name)
    if not model:
        return HttpResponseForbidden("Model not found")
    
    # Check permission
    if not check_model_permission(request.user, model, 'add'):
        return HttpResponseForbidden("You don't have permission to add this model")
    
    # Get metadata
    meta = model_registry.get_model_meta(app_label, model_name)
    
    # Create ModelForm dynamically
    form_class = modelform_factory(
        model,
        exclude=['id'] if hasattr(model, 'id') and model._meta.get_field('id').primary_key else [],
        widgets=_get_form_widgets(model)
    )
    
    if request.method == 'POST':
        form = form_class(request.POST, request.FILES)
        if form.is_valid():
            try:
                instance = form.save()
                messages.success(request, f'{meta["verbose_name"]} created successfully!')
                
                # HTMX - redirect to list
                if request.headers.get('HX-Request'):
                    response = HttpResponse()
                    response['HX-Redirect'] = f'/dashboard/model/{app_label}/{model_name}/'
                    return response
                
                return redirect('dashboard:generic_model_list', app_label=app_label, model_name=model_name)
            except Exception as e:
                messages.error(request, f'Error creating {meta["verbose_name"]}: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = form_class()
    
    context = {
        'app_label': app_label,
        'model_name': model_name,
        'model': model,
        'meta': meta,
        'form': form,
        'action': 'create',
    }
    
    # HTMX - return modal
    if request.headers.get('HX-Request'):
        return render(request, 'dashboard/generic/partials/form_modal.html', context)
    
    return render(request, 'dashboard/generic/form.html', context)


@login_required
@user_passes_test(staff_required)
def generic_model_update(request, app_label, model_name, pk):
    """Generic update view for any model"""
    model = apps.get_model(app_label, model_name)
    if not model:
        return HttpResponseForbidden("Model not found")
    
    # Check permission
    if not check_model_permission(request.user, model, 'change'):
        return HttpResponseForbidden("You don't have permission to change this model")
    
    instance = get_object_or_404(model, pk=pk)
    meta = model_registry.get_model_meta(app_label, model_name)
    
    # Create ModelForm
    form_class = modelform_factory(
        model,
        exclude=['id'] if hasattr(model, 'id') and model._meta.get_field('id').primary_key else [],
        widgets=_get_form_widgets(model)
    )
    
    if request.method == 'POST':
        form = form_class(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            try:
                instance = form.save()
                messages.success(request, f'{meta["verbose_name"]} updated successfully!')
                
                # HTMX - redirect
                if request.headers.get('HX-Request'):
                    response = HttpResponse()
                    response['HX-Redirect'] = f'/dashboard/model/{app_label}/{model_name}/'
                    return response
                
                return redirect('dashboard:generic_model_list', app_label=app_label, model_name=model_name)
            except Exception as e:
                messages.error(request, f'Error updating {meta["verbose_name"]}: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = form_class(instance=instance)
    
    context = {
        'app_label': app_label,
        'model_name': model_name,
        'model': model,
        'meta': meta,
        'form': form,
        'instance': instance,
        'action': 'update',
    }
    
    # HTMX - return modal
    if request.headers.get('HX-Request'):
        return render(request, 'dashboard/generic/partials/form_modal.html', context)
    
    return render(request, 'dashboard/generic/form.html', context)


@login_required
@user_passes_test(staff_required)
def generic_model_detail(request, app_label, model_name, pk):
    """Generic detail/view for any model"""
    model = apps.get_model(app_label, model_name)
    if not model:
        return HttpResponseForbidden("Model not found")
    
    # Check permission
    if not check_model_permission(request.user, model, 'view'):
        return HttpResponseForbidden("You don't have permission to view this model")
    
    instance = get_object_or_404(model, pk=pk)
    meta = model_registry.get_model_meta(app_label, model_name)
    
    # Get all fields for detail view
    fields_data = []
    for field in model._meta.get_fields():
        if field.auto_created and not field.concrete:
            continue
        
        if hasattr(field, 'editable') and not field.editable:
            continue
        
        fields_data.append({
            'name': field.name,
            'verbose_name': getattr(field, 'verbose_name', field.name),
            'value': get_field_display_value(instance, field.name),
        })
    
    # Get actions
    actions = get_model_actions(request.user, model)
    
    context = {
        'app_label': app_label,
        'model_name': model_name,
        'model': model,
        'meta': meta,
        'instance': instance,
        'fields_data': fields_data,
        'actions': actions,
    }
    
    # HTMX - return modal
    if request.headers.get('HX-Request'):
        return render(request, 'dashboard/generic/partials/detail_modal.html', context)
    
    return render(request, 'dashboard/generic/detail.html', context)


@login_required
@user_passes_test(staff_required)
@require_http_methods(["DELETE", "POST"])
def generic_model_delete(request, app_label, model_name, pk):
    """Generic delete view for any model"""
    model = apps.get_model(app_label, model_name)
    if not model:
        return HttpResponseForbidden("Model not found")
    
    # Check permission
    if not check_model_permission(request.user, model, 'delete'):
        return HttpResponseForbidden("You don't have permission to delete this model")
    
    instance = get_object_or_404(model, pk=pk)
    meta = model_registry.get_model_meta(app_label, model_name)
    
    try:
        instance_str = str(instance)
        instance.delete()
        messages.success(request, f'{meta["verbose_name"]} "{instance_str}" deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting {meta["verbose_name"]}: {str(e)}')
    
    # HTMX - redirect
    if request.headers.get('HX-Request'):
        response = HttpResponse()
        response['HX-Redirect'] = f'/dashboard/model/{app_label}/{model_name}/'
        return response
    
    return redirect('dashboard:generic_model_list', app_label=app_label, model_name=model_name)


@login_required
@user_passes_test(staff_required)
def generic_model_export_csv(request, app_label, model_name):
    """Export model data to CSV"""
    model = apps.get_model(app_label, model_name)
    if not model:
        return HttpResponseForbidden("Model not found")
    
    # Check permission
    if not check_model_permission(request.user, model, 'view'):
        return HttpResponseForbidden("You don't have permission to export this model")
    
    meta = model_registry.get_model_meta(app_label, model_name)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{model_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Get fields
    fields = [f for f in model._meta.get_fields() 
              if not (f.auto_created and not f.concrete) and hasattr(f, 'editable') and f.editable]
    
    # Write header
    writer.writerow([f.verbose_name for f in fields])
    
    # Write data
    for obj in model.objects.all():
        row = []
        for field in fields:
            try:
                value = getattr(obj, field.name)
                if value is None:
                    row.append('')
                elif isinstance(field, models.ManyToManyField):
                    row.append(', '.join([str(o) for o in value.all()]))
                elif isinstance(field, models.ForeignKey):
                    row.append(str(value))
                elif isinstance(field, (models.ImageField, models.FileField)):
                    row.append(value.url if value else '')
                else:
                    row.append(str(value))
            except:
                row.append('')
        writer.writerow(row)
    
    return response


@login_required
@user_passes_test(staff_required)
def generic_model_export_excel(request, app_label, model_name):
    """Export model data to Excel (requires openpyxl)"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, "Excel export requires openpyxl. Please install it.")
        return redirect('dashboard:generic_model_list', app_label=app_label, model_name=model_name)
    
    model = apps.get_model(app_label, model_name)
    if not model:
        return HttpResponseForbidden("Model not found")
    
    # Check permission
    if not check_model_permission(request.user, model, 'view'):
        return HttpResponseForbidden("You don't have permission to export this model")
    
    meta = model_registry.get_model_meta(app_label, model_name)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_name[:31]  # Excel sheet name limit
    
    # Get fields
    fields = [f for f in model._meta.get_fields() 
              if not (f.auto_created and not f.concrete) and hasattr(f, 'editable') and f.editable]
    
    # Write header
    for col_num, field in enumerate(fields, 1):
        ws.cell(row=1, column=col_num, value=field.verbose_name)
    
    # Write data
    for row_num, obj in enumerate(model.objects.all(), 2):
        for col_num, field in enumerate(fields, 1):
            try:
                value = getattr(obj, field.name)
                if value is None:
                    ws.cell(row=row_num, column=col_num, value='')
                elif isinstance(field, models.ManyToManyField):
                    ws.cell(row=row_num, column=col_num, value=', '.join([str(o) for o in value.all()]))
                elif isinstance(field, models.ForeignKey):
                    ws.cell(row=row_num, column=col_num, value=str(value))
                elif isinstance(field, (models.ImageField, models.FileField)):
                    ws.cell(row=row_num, column=col_num, value=value.url if value else '')
                else:
                    ws.cell(row=row_num, column=col_num, value=str(value))
            except:
                ws.cell(row=row_num, column=col_num, value='')
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{model_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(staff_required)
def generic_model_print(request, app_label, model_name, pk):
    """Print view for a model instance"""
    model = apps.get_model(app_label, model_name)
    if not model:
        return HttpResponseForbidden("Model not found")
    
    # Check permission
    if not check_model_permission(request.user, model, 'view'):
        return HttpResponseForbidden("You don't have permission to view this model")
    
    instance = get_object_or_404(model, pk=pk)
    meta = model_registry.get_model_meta(app_label, model_name)
    
    # Get all fields
    fields_data = []
    for field in model._meta.get_fields():
        if field.auto_created and not field.concrete:
            continue
        if hasattr(field, 'editable') and not field.editable:
            continue
        
        fields_data.append({
            'name': field.name,
            'verbose_name': getattr(field, 'verbose_name', field.name),
            'value': get_field_display_value(instance, field.name),
        })
    
    context = {
        'app_label': app_label,
        'model_name': model_name,
        'model': model,
        'meta': meta,
        'instance': instance,
        'fields_data': fields_data,
        'print_view': True,
    }
    
    return render(request, 'dashboard/generic/print.html', context)


def _get_form_widgets(model):
    """Get custom widgets for form fields"""
    from django import forms
    widgets = {}
    
    for field in model._meta.get_fields():
        if hasattr(field, 'editable') and field.editable:
            # Text fields
            if isinstance(field, models.TextField):
                widgets[field.name] = forms.Textarea(attrs={'rows': 4, 'class': 'form-control'})
            # Boolean fields
            elif isinstance(field, models.BooleanField):
                widgets[field.name] = forms.CheckboxInput(attrs={'class': 'form-check-input'})
            # Choice fields
            elif hasattr(field, 'choices') and field.choices:
                widgets[field.name] = forms.Select(attrs={'class': 'form-select'})
    
    return widgets
