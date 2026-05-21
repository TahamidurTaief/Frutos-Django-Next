"""
Import Engine - PRODUCTION GRADE
Handles CSV & Excel imports with validation
Preserves data integrity with transactions
"""

import csv
import io
from decimal import Decimal, InvalidOperation
from datetime import datetime
from django.db import transaction, models
from django.core.exceptions import ValidationError


try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ImportResult:
    """Container for import results"""
    
    def __init__(self):
        self.success_count = 0
        self.error_count = 0
        self.errors = []
        self.created_objects = []
        self.updated_objects = []
    
    def add_success(self, obj, row_num, is_update=False):
        """Record successful import"""
        self.success_count += 1
        if is_update:
            self.updated_objects.append((row_num, obj))
        else:
            self.created_objects.append((row_num, obj))
    
    def add_error(self, row_num, error_msg):
        """Record import error"""
        self.error_count += 1
        self.errors.append({'row': row_num, 'error': error_msg})
    
    def get_summary(self):
        """Get human-readable summary"""
        return {
            'total_processed': self.success_count + self.error_count,
            'success': self.success_count,
            'created': len(self.created_objects),
            'updated': len(self.updated_objects),
            'errors': self.error_count,
            'error_details': self.errors
        }


class ModelImporter:
    """Import data into Django models"""
    
    @staticmethod
    def parse_uploaded_file(file_obj):
        """Parse CSV or Excel file into rows"""
        filename = file_obj.name.lower()
        
        if filename.endswith('.csv'):
            return ModelImporter._parse_csv(file_obj)
        elif filename.endswith(('.xlsx', '.xls')):
            if not EXCEL_AVAILABLE:
                raise ImportError("openpyxl is required for Excel imports")
            return ModelImporter._parse_excel(file_obj)
        else:
            raise ValueError("Unsupported file format. Use CSV or Excel (.xlsx)")
    
    @staticmethod
    def _parse_csv(file_obj):
        """Parse CSV file"""
        # Decode file content
        content = file_obj.read().decode('utf-8-sig')  # Handle BOM
        reader = csv.DictReader(io.StringIO(content))
        
        rows = []
        for row in reader:
            # Skip empty rows
            if all(not v for v in row.values()):
                continue
            rows.append(row)
        
        return rows
    
    @staticmethod
    def _parse_excel(file_obj):
        """Parse Excel file"""
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Read data rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if all(v is None or v == '' for v in row):
                continue
            
            row_dict = {}
            for header, value in zip(headers, row):
                if header:
                    row_dict[header] = value
            
            rows.append(row_dict)
        
        return rows
    
    @staticmethod
    def import_data(model, rows, update_existing=True, skip_errors=True):
        """
        Import rows into model
        
        Args:
            model: Django model class
            rows: List of dicts with field names as keys
            update_existing: Update records if ID exists
            skip_errors: Continue on errors or rollback
        
        Returns:
            ImportResult object
        """
        result = ImportResult()
        
        # Get field mapping
        field_map = {}
        for field in model._meta.get_fields():
            if field.concrete and not field.many_to_many:
                field_map[field.name] = field
        
        # Process each row
        for row_num, row_data in enumerate(rows, start=2):  # Start at 2 (after header)
            try:
                obj_data = ModelImporter._prepare_row_data(row_data, field_map, model)
                
                # Check if updating existing
                pk_field = model._meta.pk
                pk_value = row_data.get(pk_field.name) or row_data.get('id')
                
                if pk_value and update_existing:
                    # Try to update existing
                    try:
                        obj = model.objects.get(pk=pk_value)
                        for field_name, value in obj_data.items():
                            setattr(obj, field_name, value)
                        obj.full_clean()
                        obj.save()
                        result.add_success(obj, row_num, is_update=True)
                    except model.DoesNotExist:
                        # Create new
                        obj = model(**obj_data)
                        obj.full_clean()
                        obj.save()
                        result.add_success(obj, row_num, is_update=False)
                else:
                    # Create new
                    obj = model(**obj_data)
                    obj.full_clean()
                    obj.save()
                    result.add_success(obj, row_num, is_update=False)
                
            except Exception as e:
                error_msg = str(e)
                result.add_error(row_num, error_msg)
                
                if not skip_errors:
                    raise
        
        return result
    
    @staticmethod
    def _prepare_row_data(row_data, field_map, model):
        """Convert row data to model field values"""
        obj_data = {}
        
        for field_name, raw_value in row_data.items():
            # Skip if not a valid field
            if field_name not in field_map:
                continue
            
            field = field_map[field_name]
            
            # Skip auto fields
            if isinstance(field, (models.AutoField, models.BigAutoField)):
                continue
            
            # Skip auto timestamps
            if isinstance(field, models.DateTimeField):
                if getattr(field, 'auto_now', False) or getattr(field, 'auto_now_add', False):
                    continue
            
            # Convert value
            try:
                converted_value = ModelImporter._convert_value(raw_value, field, model)
                if converted_value is not None:
                    obj_data[field_name] = converted_value
            except Exception as e:
                raise ValueError(f"Error converting field '{field_name}': {str(e)}")
        
        return obj_data
    
    @staticmethod
    def _convert_value(value, field, model):
        """Convert raw value to appropriate field type"""
        # Handle empty/null
        if value is None or value == '':
            if field.null:
                return None
            elif field.blank and isinstance(field, (models.CharField, models.TextField)):
                return ''
            else:
                # Required field with no value
                if not field.has_default():
                    raise ValueError(f"Required field '{field.name}' is empty")
                return field.get_default()
        
        # Boolean
        if isinstance(field, models.BooleanField):
            if isinstance(value, bool):
                return value
            value_str = str(value).lower().strip()
            return value_str in ('1', 'true', 'yes', 'y', 'on')
        
        # Integer
        if isinstance(field, (models.IntegerField, models.PositiveIntegerField, models.BigIntegerField)):
            try:
                return int(float(value))  # Handle "123.0" from Excel
            except (ValueError, TypeError):
                raise ValueError(f"Invalid integer: {value}")
        
        # Decimal
        if isinstance(field, models.DecimalField):
            try:
                return Decimal(str(value))
            except (InvalidOperation, ValueError):
                raise ValueError(f"Invalid decimal: {value}")
        
        # Float
        if isinstance(field, models.FloatField):
            try:
                return float(value)
            except (ValueError, TypeError):
                raise ValueError(f"Invalid float: {value}")
        
        # Date
        if isinstance(field, models.DateField) and not isinstance(field, models.DateTimeField):
            if isinstance(value, datetime):
                return value.date()
            try:
                return datetime.strptime(str(value), '%Y-%m-%d').date()
            except ValueError:
                try:
                    return datetime.strptime(str(value), '%m/%d/%Y').date()
                except ValueError:
                    raise ValueError(f"Invalid date format: {value}. Use YYYY-MM-DD")
        
        # DateTime
        if isinstance(field, models.DateTimeField):
            if isinstance(value, datetime):
                return value
            try:
                return datetime.strptime(str(value), '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    return datetime.strptime(str(value), '%Y-%m-%d')
                except ValueError:
                    raise ValueError(f"Invalid datetime format: {value}. Use YYYY-MM-DD HH:MM:SS")
        
        # ForeignKey
        if isinstance(field, models.ForeignKey):
            related_model = field.related_model
            try:
                # Try by ID
                return related_model.objects.get(pk=value)
            except (related_model.DoesNotExist, ValueError):
                # Try by string representation (name, etc.)
                try:
                    return related_model.objects.get(name=value)
                except:
                    try:
                        return related_model.objects.get(slug=value)
                    except:
                        raise ValueError(f"Cannot find {related_model.__name__} with value: {value}")
        
        # Choices
        if hasattr(field, 'choices') and field.choices:
            # Check if value is valid choice
            valid_choices = [choice[0] for choice in field.choices]
            if value in valid_choices:
                return value
            # Try to match by display name
            for choice_value, choice_display in field.choices:
                if str(value).lower() == str(choice_display).lower():
                    return choice_value
            raise ValueError(f"Invalid choice: {value}. Must be one of {valid_choices}")
        
        # String fields
        if isinstance(field, (models.CharField, models.TextField, models.EmailField, models.URLField)):
            value_str = str(value).strip()
            if isinstance(field, models.CharField) and len(value_str) > field.max_length:
                raise ValueError(f"Value too long. Max length: {field.max_length}")
            return value_str
        
        # Default - return as string
        return str(value)
    
    @staticmethod
    @transaction.atomic
    def import_with_transaction(model, rows, update_existing=True, skip_errors=False):
        """
        Import data within a transaction
        If skip_errors=False, any error will rollback all changes
        """
        return ModelImporter.import_data(model, rows, update_existing, skip_errors)
