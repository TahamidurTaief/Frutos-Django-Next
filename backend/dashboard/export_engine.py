"""
Export & Download Engine - PRODUCTION GRADE
Handles CSV & Excel exports with full relational data
Flattens relationships while preserving meaning
"""

import csv
import io
from decimal import Decimal
from datetime import date, datetime
from django.db import models
from django.http import HttpResponse
from django.core.serializers import serialize
import json


try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ModelExporter:
    """Export Django models with full relational data"""
    
    @staticmethod
    def get_exportable_fields(model):
        """Get all fields suitable for export (including relations)"""
        fields = []
        
        for field in model._meta.get_fields():
            # Skip reverse relations
            if field.auto_created and not field.concrete:
                continue
            
            # Skip many-to-many for flat export (handled separately)
            if isinstance(field, models.ManyToManyField):
                continue
            
            fields.append(field)
        
        return fields
    
    @staticmethod
    def get_field_value(obj, field):
        """Get display-ready value for a field"""
        value = getattr(obj, field.name, None)
        
        if value is None:
            return ''
        
        # Boolean
        if isinstance(field, models.BooleanField):
            return 'Yes' if value else 'No'
        
        # Choices
        if hasattr(field, 'choices') and field.choices:
            display_method = f'get_{field.name}_display'
            if hasattr(obj, display_method):
                return getattr(obj, display_method)()
        
        # DateTime
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(value, date):
            return value.strftime('%Y-%m-%d')
        
        # Decimal
        if isinstance(value, Decimal):
            return float(value)
        
        # ForeignKey - return string representation
        if isinstance(field, models.ForeignKey):
            return str(value)
        
        # File/Image - return filename or URL
        if isinstance(field, (models.FileField, models.ImageField)):
            try:
                return value.url if value else ''
            except:
                return str(value) if value else ''
        
        return str(value)
    
    @staticmethod
    def get_header_name(field):
        """Get human-readable column header"""
        return field.verbose_name.title()
    
    @staticmethod
    def export_to_csv(queryset, model):
        """Export queryset to CSV with all data"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{model._meta.model_name}_export.csv"'
        
        # UTF-8 BOM for Excel compatibility
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Get fields
        fields = ModelExporter.get_exportable_fields(model)
        
        # Write header
        headers = [ModelExporter.get_header_name(field) for field in fields]
        writer.writerow(headers)
        
        # Write data
        for obj in queryset:
            row = [ModelExporter.get_field_value(obj, field) for field in fields]
            writer.writerow(row)
        
        return response
    
    @staticmethod
    def export_to_excel(queryset, model):
        """Export queryset to Excel with formatting"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = model._meta.verbose_name_plural[:31]  # Excel limit
        
        # Get fields
        fields = ModelExporter.get_exportable_fields(model)
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = ModelExporter.get_header_name(field)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, obj in enumerate(queryset, start=2):
            for col_idx, field in enumerate(fields, start=1):
                value = ModelExporter.get_field_value(obj, field)
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
        
        # Auto-size columns
        for col_idx, field in enumerate(fields, start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(ModelExporter.get_header_name(field))
            
            # Check data length (sample first 100 rows)
            for row_idx in range(2, min(102, ws.max_row + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set width with padding
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{model._meta.model_name}_export.xlsx"'
        
        return response
    
    @staticmethod
    def export_single_record(obj, format='csv'):
        """Export a single record with all related data"""
        model = obj.__class__
        
        if format == 'excel' and EXCEL_AVAILABLE:
            return ModelExporter._export_single_excel(obj)
        else:
            return ModelExporter._export_single_csv(obj)
    
    @staticmethod
    def _export_single_csv(obj):
        """Export single record to CSV"""
        model = obj.__class__
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{model._meta.model_name}_{obj.pk}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Get fields
        fields = ModelExporter.get_exportable_fields(model)
        
        # Write as key-value pairs
        writer.writerow(['Field', 'Value'])
        for field in fields:
            writer.writerow([
                ModelExporter.get_header_name(field),
                ModelExporter.get_field_value(obj, field)
            ])
        
        # Add related data
        writer.writerow([])
        writer.writerow(['Related Data'])
        
        for related_obj in model._meta.related_objects:
            if related_obj.related_name:
                try:
                    related_queryset = getattr(obj, related_obj.related_name).all()
                    count = related_queryset.count()
                    writer.writerow([related_obj.related_name, f'{count} items'])
                except:
                    pass
        
        return response
    
    @staticmethod
    def _export_single_excel(obj):
        """Export single record to Excel with detailed view"""
        model = obj.__class__
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Details"
        
        # Styles
        header_font = Font(bold=True, size=12)
        
        # Write main fields
        fields = ModelExporter.get_exportable_fields(model)
        
        ws['A1'] = 'Field'
        ws['B1'] = 'Value'
        ws['A1'].font = header_font
        ws['B1'].font = header_font
        
        row_idx = 2
        for field in fields:
            ws.cell(row=row_idx, column=1).value = ModelExporter.get_header_name(field)
            ws.cell(row=row_idx, column=2).value = ModelExporter.get_field_value(obj, field)
            row_idx += 1
        
        # Auto-size
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        
        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{model._meta.model_name}_{obj.pk}.xlsx"'
        
        return response
    
    @staticmethod
    def generate_import_template(model, format='csv'):
        """Generate empty template for imports"""
        fields = ModelExporter.get_exportable_fields(model)
        
        if format == 'excel' and EXCEL_AVAILABLE:
            return ModelExporter._generate_excel_template(model, fields)
        else:
            return ModelExporter._generate_csv_template(model, fields)
    
    @staticmethod
    def _generate_csv_template(model, fields):
        """Generate CSV template"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{model._meta.model_name}_import_template.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Header with field names (not verbose names for import)
        headers = [field.name for field in fields if not field.primary_key or not isinstance(field, models.AutoField)]
        writer.writerow(headers)
        
        # Add example row with data types
        example_row = []
        for field in fields:
            if field.primary_key and isinstance(field, models.AutoField):
                continue
            
            if isinstance(field, models.CharField):
                example_row.append('text')
            elif isinstance(field, models.IntegerField):
                example_row.append('123')
            elif isinstance(field, models.DecimalField):
                example_row.append('99.99')
            elif isinstance(field, models.BooleanField):
                example_row.append('Yes/No')
            elif isinstance(field, models.DateField):
                example_row.append('2024-01-01')
            elif isinstance(field, models.DateTimeField):
                example_row.append('2024-01-01 12:00:00')
            elif isinstance(field, models.ForeignKey):
                example_row.append('ID or value')
            else:
                example_row.append('')
        
        writer.writerow(example_row)
        
        return response
    
    @staticmethod
    def _generate_excel_template(model, fields):
        """Generate Excel template with instructions"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Template"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        col_idx = 1
        for field in fields:
            if field.primary_key and isinstance(field, models.AutoField):
                continue
            
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field.name
            cell.fill = header_fill
            cell.font = header_font
            col_idx += 1
        
        # Instructions sheet
        ws2 = wb.create_sheet("Instructions")
        ws2['A1'] = "Import Instructions"
        ws2['A1'].font = Font(bold=True, size=14)
        
        ws2['A3'] = "1. Fill in the 'Import Template' sheet"
        ws2['A4'] = "2. Column names must match exactly"
        ws2['A5'] = "3. For Foreign Keys, use the ID or unique identifier"
        ws2['A6'] = "4. Dates: YYYY-MM-DD format"
        ws2['A7'] = "5. Booleans: Yes/No or True/False"
        
        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{model._meta.model_name}_import_template.xlsx"'
        
        return response
